import os
import csv
import openpyxl
import datetime
from openpyxl import Workbook

gse_path = "GSE-Results-2025-0209-16-53.xlsx"
subject_gse_dict = {}

its_path = "ITS-Results-2025-0210-10-29.xlsx"
subject_its_dict = {}

age_path = "Age.xlsx"
age_dict = {}

gender_path = "Gender.xlsx"
gender_dict = {}

pref_path = "Preference.xlsx"
pref_dict = {}

exp_path = "Experience.xlsx"
exp_dict = {}

subject_dp1_dict = {}

target_levels = ["B1", "B2", "B3", "B4", "B5"]

group_security = []
group_passenger = []
group_robot = []

standard_username = {
    "jiazixi": "076ecaaa4a9bbeda",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "llr": "36033e6f9b0221c3",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "飞书用户5390VU": "aac3a096b47a836f",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "Liqipei": "4f795c94300704e6",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "飞书用户8147XF": "fb54c3d25307d7a6",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "hg": "17ca5520cf72a9d9",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "飞书用户4718AD": "393f36c7f4116ce5",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "ysx": "abde714d748dd79a",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "ciel": "24481c2316cbfbba",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值

    "yizhen": "14d20be1dea6b003",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "严": "21c544db2395eccd",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "飞书用户0762JB": "ae4d0c8de33fda7f",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "周": "811b1e38e7923d87",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "飞书用户3374BA": "4bf052903d694235",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "柯柯": "f3345e299a8b4209",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "哈哈哈": "7fc12372a40f034a",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
    "蓝组qi": "321a57bc72c8be33",  # [Privacy] Participant name replaced with MD5 hash / [隐私保护] 受试者姓名已替换为MD5哈希值
}

def get_grand_grand_parent_folder():
    current_script_path = os.path.abspath(__file__)
    current_folder = os.path.dirname(current_script_path)
    parent_folder = os.path.dirname(current_folder)
    grandparent_folder = os.path.dirname(parent_folder)
    return grandparent_folder

def build_group_ref():
    global group_security
    global group_passenger
    global group_robot

    target_file_name = get_grand_grand_parent_folder() + os.sep + "Results" + os.sep + "Subject_Leader.csv"
    with open(target_file_name, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            if len(row) > 1:
                name = row[0]
                if "Passenger" in str(row[3]):
                    group_passenger.append(name)
                if "Security" in str(row[3]):
                    group_security.append(name)
                if "Robot" in str(row[3]):
                    group_robot.append(name)
    print("Passenger group total / Passenger组总计：" + str(len(group_passenger)))
    print("Security group total / Security组总计：" + str(len(group_security)))
    print("Robot group total / Robot组总计：" + str(len(group_robot)))

def name_correction(name):
    if name in standard_username.keys():
        return standard_username[name]
    else:
        return name

def build_gse_data():
    global subject_gse_dict
    workbook = openpyxl.load_workbook(gse_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "飞书用户名":
            raw_name = str(row[0])
            name = name_correction(raw_name)
            gse_value = float(row[2])
            subject_gse_dict[name] = gse_value
    workbook.close()
    # print(subject_gse_dict)

def build_its_data():
    global subject_its_dict
    workbook = openpyxl.load_workbook(its_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "飞书用户名":
            raw_name = str(row[0])
            name = name_correction(raw_name)
            its_value = float(row[2])
            subject_its_dict[name] = its_value
    workbook.close()

def build_age_data():
    global age_dict
    workbook = openpyxl.load_workbook(age_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "真实姓名":
            name = str(row[0])
            age = str(row[1])
            age_dict[name] = age

def build_gender_data():
    global gender_dict
    workbook = openpyxl.load_workbook(gender_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "真实姓名":
            name = str(row[0])
            gender = str(row[1])
            gender_dict[name] = gender

def build_pref_data():
    global pref_dict
    workbook = openpyxl.load_workbook(pref_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "真实姓名":
            name = str(row[0])
            # print(name)
            prefs = f"{str(row[1])},{str(row[2])},{str(row[3])}"
            pref_dict[name] = prefs

def build_exp_data():
    global exp_dict
    workbook = openpyxl.load_workbook(exp_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if str(row[0]) != "真实姓名":
            name = str(row[0])
            exps = f"{str(row[1])},{str(row[2])},{str(row[3])}"
            exp_dict[name] = exps

def get_dp2_result_data(subject_name, level_name):
    dp_path = get_grand_grand_parent_folder() + os.sep + "Results" + os.sep + "Subject_DP.csv"
    dp_result = "Null"
    row_id = get_level_data(level_name)
    with open(dp_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            if str(row[0]) == subject_name:
                raw_result = str(row[row_id])
                if raw_result == "Null" or raw_result == "Error":
                    dp_result = raw_result
                else:
                    if "A" in raw_result or "B" in raw_result:
                        if "A" in raw_result:
                            dp_result = "Yes"
                        else:
                            dp_result = "No"
    if dp_result == "Null":
        print(f"Name:{subject_name} Level:{level_name} DP2 Result Is Null.")
    return dp_result

def get_follow_rate(level_name):
    if level_name == "B1":
        return "0.5"
    if level_name == "B2":
        return "0.25"
    if level_name == "B3":
        return "0.75"
    if level_name == "B4":
        return "0"
    if level_name == "B5":
        return "1"

def get_level_data(level_name):
    if level_name == "A1":
        return 1
    if level_name == "A2":
        return 2
    if level_name == "B1":
        return 3
    if level_name == "B2":
        return 4
    if level_name == "B3":
        return 5
    if level_name == "B4":
        return 6
    if level_name == "B5":
        return 7

def get_leader_type(subject_name):
    global group_security
    global group_passenger
    global group_robot

    group_id = "Null"
    if subject_name in group_security:
        group_id = "Security"
    if subject_name in group_passenger:
        group_id = "Passenger"
    if subject_name in group_robot:
        group_id = "Robot"

    if group_id == "Null":
        print(f"Subject:{subject_name} missing leader type / 缺少领导类型")
    return group_id

def build_data_chain():
    out_put_file = "Data_Chain.csv"
    header = "SubjectName,GSE,ITS,Age,Gender,Familiarity,FollowTendency,SelfDecideTendency,3D,XR,MetroFrequency,Level,FollowRate,LeaderType,DP1_Result"
    with open(out_put_file, 'w', newline='', encoding='utf-8') as csvfile:
        csvfile.write(header + "\n")
        for key in subject_gse_dict.keys():
            header = key
            gse = "Null"
            if key in subject_gse_dict.keys():
                gse = subject_gse_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            its = "Null"
            if key in subject_its_dict.keys():
                its = subject_its_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            age = "Null"
            if key in age_dict.keys():
                age = age_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            gender = "Null"
            if key in gender_dict.keys():
                gender = gender_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            pref = "Null,Null,Null"
            if key in pref_dict.keys():
                pref = pref_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            exp = "Null,Null,Null"
            if key in exp_dict.keys():
                exp = exp_dict[key]
            else:
                print(f"{key} data is missing / {key} 数据缺失!")

            content = f"{gse},{its},{age},{gender},{pref},{exp}"
            counter = 0
            while counter < len(target_levels):
                level = target_levels[counter]
                dp_result = get_dp2_result_data(key, level)
                follow_rate = get_follow_rate(level)
                group_type = get_leader_type(key)
                level_add_on = f"{level},{follow_rate},{group_type},{dp_result}"
                line = f"{header},{content},{level_add_on}"
                csvfile.write(line + "\n")
                counter = counter + 1

    out_put_file_trim = "Data_Chain_Trim.csv"

    with (open(out_put_file, 'r', newline='', encoding='utf-8',errors="ignore") as infile,
          open(out_put_file_trim, 'w', newline='', encoding='utf-8', errors="ignore") as outfile):
        reader = csv.reader(infile)
        writer = csv.writer(outfile)
        counter = 0
        for row in reader:
            if "Null" not in str(row) and "Error" not in str(row):
                counter = counter + 1
                writer.writerow(row)
        print(f"Total DP2 decision data records collected / 共统计到DP2决策数据{counter - 1}组")

    os.remove(out_put_file)

if __name__ == "__main__":
    build_group_ref()
    build_gse_data()
    build_its_data()
    build_age_data()
    build_gender_data()
    build_pref_data()
    build_exp_data()
    build_data_chain()
