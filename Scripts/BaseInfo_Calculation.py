import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_real_names = []
global_subject_genders = []

global_subject_scale_data = []

feishu_names_index = 2
nick_names_index = 6
real_names_index = 4
gender_index = 5

scale_data_min_index = 7
scale_data_max_index = 14

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "基本情况问卷" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">>已找到目标文件")
        print(global_file_path)
    else:
        print(">>未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_real_names
    global global_subject_genders

    global global_subject_scale_data

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=real_names_index, max_col=real_names_index, values_only=True):
        global_subject_real_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=gender_index, max_col=gender_index, values_only=True):
        global_subject_genders.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">>飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">>昵称总计：" + str(len(global_subject_nick_names)))
    print(">>量表数据总计：" + str(len(global_subject_scale_data)))
    print("\n" + "飞书用户名")
    print(global_subject_feishu_names)
    print("\n" + "昵称")
    print(global_subject_nick_names)
    print("\n" + "真实姓名")
    print(global_subject_real_names)
    print("\n" + "性别")
    print(global_subject_genders)
    print("\n" + "量表数据")
    print(global_subject_scale_data)

    print(">>完成读取量表数据")

def output_age():
    output_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    output_path = output_path + "Results" + os.sep + "BaseInfo" + os.sep + "Age-" + timestamp + ".xlsx"
    print("目标输出路径已确认：" + output_path)

    subject_ages = []
    index = 0
    while index < len(global_subject_scale_data):
        datas = global_subject_scale_data[index].split(",")
        age = datas[0]
        age = age.replace("岁", "")
        subject_ages.append(age)
        index = index + 1

    print(">>年龄总计：" + str(len(subject_ages)))

    wb = Workbook()
    ws = wb.active
    output_base_info(ws)

    ws['E1'] = '年龄'
    for row_index, age in enumerate(subject_ages, start=2):
        ws.cell(row=row_index, column=5, value=age)

    wb.save(output_path)

def output_location():
    output_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    output_path = output_path + "Results" + os.sep + "BaseInfo" + os.sep + "Location-" + timestamp + ".xlsx"
    print("目标输出路径已确认：" + output_path)

    subject_pos = []
    index = 0
    while index < len(global_subject_scale_data):
        datas = global_subject_scale_data[index].split(",")
        pos = datas[3]
        subject_pos.append(pos)
        index = index + 1

    print(">>位置总计：" + str(len(subject_pos)))

    wb = Workbook()
    ws = wb.active
    output_base_info(ws)

    ws['E1'] = '所在地'
    for row_index, pos in enumerate(subject_pos, start=2):
        ws.cell(row=row_index, column=5, value=pos)

    wb.save(output_path)

def output_experience():
    output_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    output_path = output_path + "Results" + os.sep + "BaseInfo" + os.sep + "Experience-" + timestamp + ".xlsx"
    print("目标输出路径已确认：" + output_path)

    subject_3d_experience = []
    subject_vr_experience = []
    subject_metro_experience = []

    index = 0
    while index < len(global_subject_scale_data):
        datas = global_subject_scale_data[index].split(",")
        subject_3d_experience.append(get_experience_value(datas[1]))
        subject_vr_experience.append(get_experience_value(datas[2]))
        subject_metro_experience.append(get_experience_value(datas[4]))
        index = index + 1

    wb = Workbook()
    ws = wb.active
    output_base_info(ws)

    ws['E1'] = '3D游戏经验'
    ws['F1'] = 'XR使用经验'
    ws['G1'] = '城轨交通使用频率'

    for row_index, _3d in enumerate(subject_3d_experience, start=2):
        ws.cell(row=row_index, column=5, value=_3d)

    for row_index, _vr in enumerate(subject_vr_experience, start=2):
        ws.cell(row=row_index, column=6, value=_vr)

    for row_index, _metro in enumerate(subject_metro_experience, start=2):
        ws.cell(row=row_index, column=7, value=_metro)

    wb.save(output_path)

def output_preference():
    output_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    output_path = output_path + "Results" + os.sep + "BaseInfo" + os.sep + "Preference-" + timestamp + ".xlsx"
    print("目标输出路径已确认：" + output_path)

    subject_space_sense = []
    subject_follow_trend = []
    subject_self_judge = []

    index = 0
    while index < len(global_subject_scale_data):
        datas = global_subject_scale_data[index].split(",")
        subject_space_sense.append(datas[5])
        subject_follow_trend.append(datas[6])
        subject_self_judge.append(datas[7])
        index = index + 1

    wb = Workbook()
    ws = wb.active
    output_base_info(ws)

    ws['E1'] = '空间布局熟悉程度'
    ws['F1'] = '跟随倾向'
    ws['G1'] = '自行决策倾向'

    for row_index, space_sense in enumerate(subject_space_sense, start=2):
        ws.cell(row=row_index, column=5, value=space_sense)

    for row_index, follow in enumerate(subject_follow_trend, start=2):
        ws.cell(row=row_index, column=6, value=follow)

    for row_index, self in enumerate(subject_self_judge, start=2):
        ws.cell(row=row_index, column=7, value=self)

    wb.save(output_path)

def output_base_info(ws):
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_real_names
    global global_subject_genders

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = '真实姓名'
    ws['D1'] = '性别'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, real_name in enumerate(global_subject_real_names, start=2):
        ws.cell(row=row_index, column=3, value=real_name)

    for row_index, gender in enumerate(global_subject_genders, start=2):
        ws.cell(row=row_index, column=4, value=gender)

def output():
    confirm_path()
    output_age()
    output_location()
    output_experience()
    output_preference()

def confirm_path():
    output_path = global_current_script_path.replace("Scripts\BaseInfo_Calculation.py", "")
    output_dir = output_path + "Results" + os.sep + "BaseInfo"
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    else:
        clear_path(output_dir)

def clear_path(folder_path):
    try:
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"已删除文件: {file_path}")
        print(f"文件夹 {folder_path} 中的所有文件已清除。")
    except FileNotFoundError:
        print(f"指定的文件夹 {folder_path} 不存在。")
    except PermissionError:
        print(f"没有权限删除文件夹 {folder_path} 中的文件。")

def get_experience_value(str_value):
    if str_value == "我没有玩过3D游戏":
        return "0"
    if str_value == "是的，我玩过3D游戏":
        return "1"

    if str_value == "我没有使用过VR/AR/MR设备":
        return "0"
    if str_value == "是的，我曾经使用过VR/AR/MR设备":
        return "1"

    if str_value == "从没乘坐过":
        return "0"
    if str_value == "偶尔":
        return "1"
    if str_value == "二~四周一次":
        return "2"
    if str_value == "每周一次":
        return "3"
    if str_value == "每周乘坐多次，但不至于每天都乘坐":
        return "4"
    if str_value == "每天一次":
        return "5"
    if str_value == "每天多次":
        return "6"
    return "Null"

if __name__ == "__main__":
    find_file()
    read()
    output()
