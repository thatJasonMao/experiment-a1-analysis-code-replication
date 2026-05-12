import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_real_names = []
global_subject_genders = []

global_subject_scale_data = []

global_record_names = {}

feishu_names_index = 2
nick_names_index = 6
real_names_index = 4
gender_index = 5

scale_data_min_index = 7
scale_data_max_index = 14

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\Build_Global_Reference.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "基本情况问卷" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">>已找到目标文件")
        print(global_file_path)
    else:
        print(">>未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_real_names
    global global_subject_genders

    global global_subject_scale_data

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=real_names_index, max_col=real_names_index, values_only=True):
        global_subject_real_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=gender_index, max_col=gender_index, values_only=True):
        global_subject_genders.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">>飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">>昵称总计：" + str(len(global_subject_nick_names)))
    print(">>量表数据总计：" + str(len(global_subject_scale_data)))
    print("\n" + "飞书用户名")
    print(global_subject_feishu_names)
    print("\n" + "昵称")
    print(global_subject_nick_names)
    print("\n" + "真实姓名")
    print(global_subject_real_names)
    print("\n" + "性别")
    print(global_subject_genders)
    print("\n" + "量表数据")
    print(global_subject_scale_data)

    print(">>完成读取量表数据")

    global global_record_names

    target_folder_name_1 = "受试者采集数据_成都"
    target_folder_name_2 = "受试者采集数据_深圳"
    target_folder_name_3 = "受试者采集数据_长春"

    upper_folder_path = os.path.abspath(__file__).replace("Scripts\Build_Global_Reference.py", "")
    target_path_1 = upper_folder_path + target_folder_name_1
    target_path_2 = upper_folder_path + target_folder_name_2
    target_path_3 = upper_folder_path + target_folder_name_3

    traverse_folder(target_path_1)
    traverse_folder(target_path_2)
    traverse_folder(target_path_3)

    print(">>完成映射文件路径，索引总计：" + str(len(global_record_names.keys())))
    # print(global_record_names)

def traverse_folder(str_path):
    global global_record_names

    for item in os.listdir(str_path):
        item_path = os.path.join(str_path, item)
        if os.path.isdir(item_path):
            datas = item.split('_')
            name = datas[1]
            global_record_names[name] = item_path

def clear_path(str_path):
    if os.path.exists(str_path):
        os.remove(str_path)

def output_info():
    output_path = global_current_script_path.replace("Scripts\Build_Global_Reference.py", "")
    output_path = output_path + "Subject_Data_Global_Reference" + ".xlsx"
    clear_path(output_path)
    print("目标输出路径已确认：" + output_path)

    subject_ages = []
    index = 0
    while index < len(global_subject_scale_data):
        datas = global_subject_scale_data[index].split(",")
        age = datas[0]
        age = age.replace("岁", "")
        subject_ages.append(age)
        index = index + 1

    print(">>年龄总计：" + str(len(subject_ages)))

    wb = Workbook()
    ws = wb.active
    output_base_info(ws)

    ws['E1'] = '年龄'
    for row_index, age in enumerate(subject_ages, start=2):
        ws.cell(row=row_index, column=5, value=age)

    global global_subject_real_names

    folder_path_correction = []
    index_existing_table = 0
    while index_existing_table < len(global_subject_scale_data):
        current_name = global_subject_real_names[index_existing_table]
        if current_name in global_record_names.keys():
            folder_path_correction.append(global_record_names[current_name])
        else:
            folder_path_correction.append("Null")
            print(">>映射出错 受试者在数据表中存在，数据索引为空！ ID：" + current_name)
        index_existing_table = index_existing_table + 1

    ws['F1'] = '数据路径'
    for row_index, path in enumerate(folder_path_correction, start=2):
        ws.cell(row=row_index, column=6, value=path)

    ids = []
    counter = 0
    while counter < len(global_subject_scale_data):
        ids.append(str(counter + 1))
        counter = counter + 1

    ws['G1'] = '全局索引'
    for row_index, id in enumerate(ids, start=2):
        ws.cell(row=row_index, column=7, value=id)

    wb.save(output_path)

def output_base_info(ws):
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_real_names
    global global_subject_genders

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = '真实姓名'
    ws['D1'] = '性别'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, real_name in enumerate(global_subject_real_names, start=2):
        ws.cell(row=row_index, column=3, value=real_name)

    for row_index, gender in enumerate(global_subject_genders, start=2):
        ws.cell(row=row_index, column=4, value=gender)

if __name__ == "__main__":
    find_file()
    read()
    output_info()
