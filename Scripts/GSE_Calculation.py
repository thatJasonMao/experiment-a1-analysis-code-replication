import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""
global_output_path = ""

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_scale_data = []
global_subject_gse_values = []

feishu_names_index = 2
nick_names_index = 14

scale_data_min_index = 4
scale_data_max_index = 13

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\GSE_Calculation.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "自我效能感问卷" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">> Target file found / 已找到目标文件")
        print(global_file_path)
    else:
        print(">> Target file NOT found / 未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_scale_data

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">> Feishu users total / 飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">> Nickname total / 昵称总计：" + str(len(global_subject_nick_names)))
    print(">> Scale data total / 量表数据总计：" + str(len(global_subject_scale_data)))
    print("
" + "Feishu user name / 飞书用户名")
    print(global_subject_feishu_names)
    print("
" + "Nickname / 昵称")
    print(global_subject_nick_names)
    print("
" + "Scale data / 量表数据")
    print(global_subject_scale_data)

    print(">> Scale data reading complete / 完成读取量表数据")

def calculate():
    global global_subject_gse_values
    index_gse = 0
    while index_gse < len(global_subject_scale_data):
        gse_datas = global_subject_scale_data[index_gse].split(",")
        sum_gse = 0
        index_ges_item = 0
        while index_ges_item < len(gse_datas):
            sum_gse = sum_gse + int(gse_datas[index_ges_item])
            index_ges_item = index_ges_item + 1
        gse_value = sum_gse / len(gse_datas)
        print("GSE量度：" + str(gse_value))
        global_subject_gse_values.append(gse_value)
        if len(gse_datas) != 10:
            print(">>Error: 数据解析出错")
        index_gse = index_gse + 1

def output():
    global global_output_path
    global global_current_script_path

    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_gse_values

    global_output_path = global_current_script_path.replace("Scripts\GSE_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    global_output_path = global_output_path + "Results" + os.sep + "GSE-Results-" + timestamp + ".xlsx"
    check_existing(global_output_path)
    print("Output path confirmed / 目标输出路径已确认：" + global_output_path)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = 'GSE-自我效能感'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, gse in enumerate(global_subject_gse_values, start=2):
        ws.cell(row=row_index, column=3, value=gse)

    wb.save(global_output_path)

def check_existing(path):
    if os.path.exists(path):
        os.remove(path)
        print(">> Duplicate file found and removed / >>重名文件存在 已移除")
    else:
        print(">> No duplicate files found / >>未发现重名文件")

if __name__ == "__main__":
    find_file()
    read()
    calculate()
    output()
