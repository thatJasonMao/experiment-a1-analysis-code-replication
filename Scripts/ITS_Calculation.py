import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""
global_output_path = ""

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_scale_data = []
global_subject_its_values = []

feishu_names_index = 2
nick_names_index = 4

scale_data_min_index = 5
scale_data_max_index = 29

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\ITS_Calculation.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "人际信任问卷" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">>已找到目标文件")
        print(global_file_path)
    else:
        print(">>未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_scale_data

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">>飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">>昵称总计：" + str(len(global_subject_nick_names)))
    print(">>量表数据总计：" + str(len(global_subject_scale_data)))
    print("\n" + "飞书用户名")
    print(global_subject_feishu_names)
    print("\n" + "昵称")
    print(global_subject_nick_names)
    print("\n" + "量表数据")
    print(global_subject_scale_data)

    print(">>完成读取量表数据")

def calculate():
    global global_subject_its_values
    index_its = 0
    while index_its < len(global_subject_scale_data):
        its_datas = global_subject_scale_data[index_its].split(",")
        sum_its = 0
        index_its_item = 0
        while index_its_item < len(its_datas):
            sum_its = sum_its + get_scale(its_datas[index_its_item])
            index_its_item = index_its_item + 1
        print("ITS量度：" + str(sum_its))
        global_subject_its_values.append(sum_its)
        if len(its_datas) != 25:
            print(">>Error: 数据解析出错")
        index_its = index_its + 1
    print(">>ITS数据总计：" + str(len(global_subject_its_values)))

def output():
    global global_output_path
    global global_current_script_path

    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_its_values

    global_output_path = global_current_script_path.replace("Scripts\ITS_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    global_output_path = global_output_path + "Results" + os.sep + "ITS-Results-" + timestamp + ".xlsx"
    check_existing(global_output_path)
    print("目标输出路径已确认：" + global_output_path)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = 'ITS-人际信任感'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, gse in enumerate(global_subject_its_values, start=2):
        ws.cell(row=row_index, column=3, value=gse)

    wb.save(global_output_path)

def check_existing(path):
    if os.path.exists(path):
        os.remove(path)
        print(">>重名文件存在 已移除")
    else:
        print(">>未发现重名文件")

def get_scale(str_value):
    if str_value == "5":
        return 1
    if str_value == "4":
        return 2
    if str_value == "3":
        return 3
    if str_value == "2":
        return 4
    if str_value == "1":
        return 5

if __name__ == "__main__":
    find_file()
    read()
    calculate()
    output()
