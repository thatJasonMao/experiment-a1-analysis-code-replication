import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""
global_output_path = ""

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_scale_data = []

global_pa_values = []
global_na_values = []

feishu_names_index = 2
nick_names_index = 22

scale_data_min_index = 4
scale_data_max_index = 21

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\Pre_PANAS_Calculation.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "积极消极情绪问卷_实验前" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">> Target file found / 已找到目标文件")
        print(global_file_path)
    else:
        print(">> Target file NOT found / 未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">> Feishu users total / 飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">> Nickname total / 昵称总计：" + str(len(global_subject_nick_names)))
    print(">> Scale data total / 量表数据总计：" + str(len(global_subject_scale_data)))
    print("
" + "Feishu user name / 飞书用户名")
    print(global_subject_feishu_names)
    print("
" + "Nickname / 昵称")
    print(global_subject_nick_names)
    print("
" + "Scale data / 量表数据")
    print(global_subject_scale_data)

    print(">> Scale data reading complete / 完成读取量表数据")

def calculate():
    global global_pa_values
    global global_na_values

    index_pa = 0
    while index_pa < len(global_subject_scale_data):
        pa_datas = global_subject_scale_data[index_pa].split(",")
        sum_pa = (int(pa_datas[0]) + int(pa_datas[3]) + int(pa_datas[5]) + int(pa_datas[6]) + int(pa_datas[10]) + int(pa_datas[11])
                  + int(pa_datas[13]) + int(pa_datas[14]) + int(pa_datas[17]))
        global_pa_values.append(sum_pa / 9)
        print(">> PA value / >>pa值：" + str(sum_pa / 9))
        index_pa = index_pa + 1
    print(">> PA total count / >>pa值总计：" + str(len(global_pa_values)))

    index_na = 0
    while index_na < len(global_subject_scale_data):
        na_datas = global_subject_scale_data[index_na].split(",")
        sum_na = (int(na_datas[1]) + int(na_datas[2]) + int(na_datas[4]) + int(na_datas[7]) + int(na_datas[8]) + int(na_datas[9])
                  + int(na_datas[12]) + int(na_datas[15]) + int(na_datas[16]))
        global_na_values.append(sum_na / 9)
        print(">> NA value / >>na值：" + str(sum_na / 9))
        index_na = index_na + 1
    print(">> NA total count / >>na值总计：" + str(len(global_na_values)))

def output():
    global global_output_path
    global global_current_script_path

    global global_subject_feishu_names
    global global_subject_nick_names
    global global_pa_values
    global global_na_values

    global_output_path = global_current_script_path.replace("Scripts\Pre_PANAS_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    global_output_path = global_output_path + "Results" + os.sep + "Pre-PANAS-Results-" + timestamp + ".xlsx"
    check_existing(global_output_path)
    print("Output path confirmed / 目标输出路径已确认：" + global_output_path)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = 'PA-积极情绪值'
    ws['D1'] = 'NA-消极情绪值'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, pa_value in enumerate(global_pa_values, start=2):
        ws.cell(row=row_index, column=3, value=pa_value)

    for row_index, na_value in enumerate(global_na_values, start=2):
        ws.cell(row=row_index, column=4, value=na_value)

    wb.save(global_output_path)

def check_existing(path):
    if os.path.exists(path):
        os.remove(path)
        print(">> Duplicate file found and removed / >>重名文件存在 已移除")
    else:
        print(">> No duplicate files found / >>未发现重名文件")

if __name__ == "__main__":
    find_file()
    read()
    calculate()
    output()
