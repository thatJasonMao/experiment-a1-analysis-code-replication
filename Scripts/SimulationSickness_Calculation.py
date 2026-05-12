import os
import openpyxl
import datetime
from openpyxl import Workbook

global_current_script_path = ""
global_file_path = ""
global_output_path = ""

n_const = 9.54
o_const = 7.58
d_const = 13.92
sum_const = 3.74

global_subject_feishu_names = []
global_subject_nick_names = []
global_subject_scale_data = []

global_subject_ssq_value = []
global_subject_n_value = []
global_subject_o_value = []
global_subject_d_value = []

feishu_names_index = 2
nick_names_index = 20

scale_data_min_index = 4
scale_data_max_index = 19

def find_file():
    global global_current_script_path
    global_current_script_path = os.path.abspath(__file__)
    global global_file_path
    global_file_path = global_current_script_path.replace("Scripts\SimulationSickness_Calculation.py", "")
    global_file_path = global_file_path + "问卷数据" + os.sep + "仿真不适问卷" + os.sep
    global_file_path = global_file_path + "问卷.xlsx"
    if os.path.exists(global_file_path):
        print(">> Target file found / 已找到目标文件")
        print(global_file_path)
    else:
        print(">> Target file NOT found / 未找到目标文件")

def read():
    global global_subject_feishu_names
    global global_subject_nick_names

    workbook = openpyxl.load_workbook(global_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=feishu_names_index, max_col=feishu_names_index, values_only=True):
        global_subject_feishu_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=nick_names_index, max_col=nick_names_index, values_only=True):
        global_subject_nick_names.append(row[0])

    for row in sheet.iter_rows(min_row=2, min_col=scale_data_min_index, max_col=scale_data_max_index, values_only=True):
        row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
        global_subject_scale_data.append(row_str)

    workbook.close()

    print(">> Feishu users total / 飞书用户名称总计：" + str(len(global_subject_feishu_names)))
    print(">> Nickname total / 昵称总计：" + str(len(global_subject_nick_names)))
    print(">> Scale data total / 量表数据总计：" + str(len(global_subject_scale_data)))
    print("
" + "Feishu user name / 飞书用户名")
    print(global_subject_feishu_names)
    print("
" + "Nickname / 昵称")
    print(global_subject_nick_names)
    print("
" + "Scale data / 量表数据")
    print(global_subject_scale_data)

    print(">> Scale data reading complete / 完成读取量表数据")

def get_scale(str_value):
    if str_value == "无":
        return 0
    if str_value == "有点":
        return 1
    if str_value == "适中":
        return 2
    if str_value == "很强":
        return 3

def calculate():
    global global_subject_ssq_value
    global global_subject_n_value
    global global_subject_o_value
    global global_subject_d_value

    index_scale = 0
    while index_scale < len(global_subject_scale_data):
        ssq_scale_datas_int = []
        ssq_scale_datas_str = global_subject_scale_data[index_scale].split(",")
        index_scale_datas_str = 0
        while index_scale_datas_str < len(ssq_scale_datas_str):
            ssq_scale_datas_int.append(get_scale(ssq_scale_datas_str[index_scale_datas_str]))
            index_scale_datas_str = index_scale_datas_str + 1
        print(">> SSQ full data / >>SSQ全量数据")
        print(ssq_scale_datas_str)
        print(">> SSQ converted scale / >>SSQ转换后尺度")
        print(ssq_scale_datas_int)
        index_scale = index_scale + 1

        n_scale = (ssq_scale_datas_int[0] + ssq_scale_datas_int[6] + ssq_scale_datas_int[8] + ssq_scale_datas_int[14]
                   + ssq_scale_datas_int[15])

        o_scale = ssq_scale_datas_int[2] + ssq_scale_datas_int[3] + ssq_scale_datas_int[4] + ssq_scale_datas_int[10]

        d_scale = ssq_scale_datas_int[11] + ssq_scale_datas_int[12] + ssq_scale_datas_int[13]

        print(">> Current N scale sum (Nausea) / >>当前N尺度累加（恶心） " + str(n_scale) + " 当前O尺度累加（眼动障碍） " + str(o_scale) + " 当前D尺度累加（方向障碍） " + str(d_scale))
        ssq_value = (n_scale + o_scale + d_scale) * sum_const
        print(">> Current SSQ total / >>当前SSQ总值：" + str(ssq_value))
        global_subject_ssq_value.append(ssq_value)

        n_value = n_scale * n_const
        o_value = o_scale * o_const
        d_value = d_scale * d_const
        print(">> N value (Nausea) / >>N值（恶心） " + str(n_value) + " O值（眼动障碍） " + str(o_value) + " D值（方向障碍） " + str(d_value))
        print("\n")
        global_subject_n_value.append(n_value)
        global_subject_o_value.append(o_value)
        global_subject_d_value.append(d_value)

def output():
    global global_output_path
    global global_current_script_path

    global global_subject_feishu_names
    global global_subject_nick_names
    global global_subject_ssq_value
    global global_subject_n_value
    global global_subject_o_value
    global global_subject_d_value

    global_output_path = global_current_script_path.replace("Scripts\SimulationSickness_Calculation.py", "")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m%d-%H-%M")
    global_output_path = global_output_path + "Results" + os.sep + "SSQ-Results-" + timestamp + ".xlsx"
    check_existing(global_output_path)
    print("Output path confirmed / 目标输出路径已确认：" + global_output_path)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '飞书用户名'
    ws['B1'] = '昵称'
    ws['C1'] = 'SSQ-仿真不适'
    ws['D1'] = 'N-恶心尺度'
    ws['E1'] = 'O-眼动障碍尺度'
    ws['F1'] = 'D-方向障碍尺度'

    for row_index, feishu_name in enumerate(global_subject_feishu_names, start=2):
        ws.cell(row=row_index, column=1, value=feishu_name)

    for row_index, nick_name in enumerate(global_subject_nick_names, start=2):
        ws.cell(row=row_index, column=2, value=nick_name)

    for row_index, ssq in enumerate(global_subject_ssq_value, start=2):
        ws.cell(row=row_index, column=3, value=ssq)

    for row_index, n in enumerate(global_subject_n_value, start=2):
        ws.cell(row=row_index, column=4, value=n)

    for row_index, o in enumerate(global_subject_o_value, start=2):
        ws.cell(row=row_index, column=5, value=o)

    for row_index, d in enumerate(global_subject_d_value, start=2):
        ws.cell(row=row_index, column=6, value=d)

    wb.save(global_output_path)

def check_existing(path):
    if os.path.exists(path):
        os.remove(path)
        print(">> Duplicate file found and removed / >>重名文件存在 已移除")
    else:
        print(">> No duplicate files found / >>未发现重名文件")

if __name__ == "__main__":
    find_file()
    read()
    calculate()
    output()
